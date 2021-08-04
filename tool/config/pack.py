# -*- coding: UTF-8 -*-
import sys
import os
import openpyxl
import time

reload(sys)
sys.setdefaultencoding('utf-8')

def packExcel():
    tm = time.time()
    nums = 0
    tsStr = 'export interface ICfg {\n\t/** 唯一id */\n\tid: number;\n}\n\n'
    jsonStr = '{'
    modelStr = writeModelStart()

    excelUrl = getPath('excelDir')
    excelList = os.listdir(excelUrl)
    for excel in excelList:
        # 排除隐藏的文件
        if excel[0] == '.':
            printCN('过滤隐藏的文件：%s' % excel)
            continue
        # 排除打开的文件
        if excel[0:2] == '~$':
            printCN('过滤打开的文件：%s' % excel)
            continue
        # 排除不是表格的文件
        fileType = excel.split('.')[1]
        if fileType != 'xlsx' and fileType != 'xls':
            printCN('过滤不是excel的文件：%s' % excel)
            continue
        url = excelUrl + excel
        wb = openpyxl.load_workbook(url)
        for sheetName in wb.sheetnames:
            sheet = wb[sheetName]
            # 排除空表格
            if sheet.cell(2, 1).value != 'id':
                printCN('过滤空表格：%s' % sheetName)
                continue

            # 编写TS数据
            tsStr += writeTs(sheet, sheetName, excel)
            # 编写Json数据
            jsonData = writeJson(sheet, sheetName)
            jsonStr += jsonData
            jsonData = '{%s}' % jsonData.strip(jsonData[-1])
            writeToFile(getPath('jsonDir') + sheetName + '.json', jsonData)
            # 编写model数据
            modelStr += writeModel(sheetName)
            print(sheetName)
            nums = nums + 1

    writeToFile(getPath('interfaceDir'), tsStr)
    printCN('写入TS完成')

    jsonStr = jsonStr.strip(jsonStr[-1]) + '}'
    writeToFile(getPath('dataDir'), jsonStr)
    printCN('写入json完成')

    modelStr += writeModelEnd()
    writeToFile(getPath('modelDir'), modelStr)

    cd = time.time() - tm
    printCN('花费的时间：%s' % cd)
    printCN('表格数量：%s' % nums)

# 编写TS数据
def writeTs(sheet, sheetName, filename):
    tsStr = '/** %s */\n' % filename
    tsStr += 'export interface I%sCfg extends ICfg {\n' % getClassName(sheetName)
    # 描述
    row1 = sheet[1]
    # 变量名
    row2 = sheet[2]
    # 变量类型
    row3 = sheet[3]
    # 是否客户端用到
    row4 = sheet[4]
    num = len(row2)
    for i in range(num):
        if row2[i].value is None:
            break
        if row4[i].value != 'both' and row4[i].value != 'client':
            continue
        if row2[i].value == 'id':
            continue
        desc = row1[i].value
        var = row2[i].value
        varType = getType(row3[i].value)
        tsStr += '\t/** %s */\n' % desc
        tsStr += '\t%s: %s;\n' % (var, varType)
    tsStr += '}\n\n'
    return tsStr

def writeModelStart():
    tsStr = '/** 工具生成，请勿手动修改 */\n'
    tsStr += 'import { BaseIns } from "../base/BaseIns";\n\n'
    tsStr += 'export class ConfigModel extends BaseIns {\n'
    tsStr += '\tpublic static get ins(): ConfigModel {\n'
    tsStr += '\t\treturn super.ins as ConfigModel;\n'
    tsStr += '\t}\n\n'
    tsStr += '\tprivate _cfgDic = {};\n\n'
    return tsStr

def writeModel(sheetName):
    tsStr = ''
    funName = getFunName(sheetName)
    name = getClassName(sheetName)
    tsStr += '\tpublic get %s() {\n' % funName
    tsStr += '\t\treturn this._cfgDic["%s"];\n' % name
    tsStr += '\t}\n\n'
    return tsStr

def writeModelEnd():
    tsStr = '\tpublic parseCfg(data: any): void {\n'
    tsStr += '\t\tfor (let k in data) {\n'
    tsStr += '\t\t\tthis._cfgDic[k] = data[k];\n'
    tsStr += '\t\t}\n\t}\n}'
    return tsStr

# 编写json数据
def writeJson(sheet, sheetName):
    jsonStr = '"%s": {' % getClassName(sheetName)
    maxRow = sheet.max_row
    maxCol = sheet.max_column
    for i in range(5, maxRow + 1):
        # 如果该行第一个数据已经是空了，那就直接跳出循环
        if sheet[i][0].value is None:
            break
        uid = sheet.cell(i, 1).value
        t = '"%s": {' % uid
        for j in range(1, maxCol + 1):
            client = sheet.cell(4, j).value
            if client != 'both' and client != 'client':
                continue
            key = '%s' % sheet.cell(2, j).value
            val = sheet.cell(i, j).value
            t += '"%s": "%s",' % (key, val)
        # 去掉字符串最后一个逗号
        t = t.strip(t[-1])
        jsonStr += t + '},'
    jsonStr = jsonStr.strip(jsonStr[-1]) + '},'
    return jsonStr

def writeToFile(url, fileData):
    with open(url, 'w') as f:
        f.write(fileData)

# 获取json文件里的路径
def getPath(pathName):
    with open('./config.json', 'r') as f:
        content = f.read()
        # 将字符串转换成字典格式
        data = eval(content)
    return data[pathName]

# 获取类名
def getClassName(sheetName):
    name = ''
    arr = sheetName.split('_')
    # 将首字母都转换成大写，其它改为小写
    # mp = list(map(lambda x: x[0].upper() + x[1:].lower(), arr))
    mp = list(map(lambda x: x.capitalize(), arr))
    for v in mp:
        name += v
    return name

def getFunName(sheetName):
    name = ''
    arr = sheetName.split('_')
    for i in range(len(arr)):
        if i == 0:
            name += arr[i].lower()
            continue
        s = arr[i].capitalize()
        name += s
    return name

def getType(value):
    varType = ''
    if value == 'int':
        varType = 'number'
    elif value == 'array':
        varType = 'Array<any>'
    elif value == 'object':
        varType = 'any'
    elif value == 'string':
        varType = 'string'
    else:
        varType = 'undefined'
    return varType

def printCN(desc):
    print(desc.encode('gb18030'))


if __name__ == "__main__":
    packExcel()